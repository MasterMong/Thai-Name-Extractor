import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from docx import Document
from collections import Counter
import re
from openpyxl import Workbook
from datetime import datetime
from PyPDF2 import PdfReader

class NameExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Thai Name Extractor")
        self.root.geometry("800x600")

        # Set Thai font
        self.thai_font = ('TH Sarabun New', 14)
        
        # Store all names for filtering and sorting state
        self.all_names = []
        self.sort_reverse = {'Name': False, 'Count': True}
        
        # Create button frame
        button_frame = tk.Frame(root)
        button_frame.pack(pady=10)
        
        # Move select button to button frame
        self.select_btn = tk.Button(button_frame, text="เลือกไฟล์ DOCX, PDF", command=self.select_file, font=self.thai_font)
        self.select_btn.pack(side=tk.LEFT, padx=5)
        
        # Add export button
        self.export_btn = tk.Button(button_frame, text="ส่งออก Excel", command=self.export_to_excel, font=self.thai_font)
        self.export_btn.pack(side=tk.LEFT, padx=5)

        # Add search frame
        search_frame = tk.Frame(root)
        search_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Add search entry
        tk.Label(search_frame, text="ค้นหา:", font=self.thai_font).pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_names)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=self.thai_font)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        # Create treeview with Thai font
        self.tree = ttk.Treeview(root, columns=('Name', 'Count'), show='headings')
        self.tree.heading('Name', text='ชื่อ ↕', command=lambda: self.sort_treeview('Name'))
        self.tree.heading('Count', text='จำนวน ↕', command=lambda: self.sort_treeview('Count'))
        self.tree.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
        
        # Configure column widths
        self.tree.column('Name', width=600)
        self.tree.column('Count', width=100)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(root, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

    def sort_treeview(self, column):
        items = [(self.tree.set(item, column), item) for item in self.tree.get_children('')]
        
        # Convert count to integer for numeric sorting if sorting by Count
        if column == 'Count':
            items = [(int(value), item) for value, item in items]
            
        # Toggle sort direction
        self.sort_reverse[column] = not self.sort_reverse[column]
        items.sort(reverse=self.sort_reverse[column])
        
        # Update column headers to show sort direction
        self.tree.heading('Name', text='ชื่อ ' + ('↓' if self.sort_reverse['Name'] else '↑') 
                         if column == 'Name' else 'ชื่อ ↕')
        self.tree.heading('Count', text='จำนวน ' + ('↓' if self.sort_reverse['Count'] else '↑') 
                         if column == 'Count' else 'จำนวน ↕')
        
        # Rearrange items in sorted positions
        for index, (val, item) in enumerate(items):
            self.tree.move(item, '', index)

    def filter_names(self, *args):
        # Clear current items
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        search_term = self.search_var.get().lower()
        
        # Filter and display matching items
        filtered_names = [(name, count) for name, count in self.all_names 
                         if search_term in name.lower()]
        
        for name, count in filtered_names:
            self.tree.insert('', tk.END, values=(name, count))

    def select_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[
                ("Document Files", "*.docx *.pdf"),
                ("Word Documents", "*.docx"),
                ("PDF Files", "*.pdf")
            ]
        )
        if file_path:
            self.process_file(file_path)

    def clean_thai_name(self, name):
        """Clean name by normalizing spaces and keeping only title and first two words"""
        # Replace multiple spaces with single space and strip
        name = re.sub(r'\s+', ' ', name).strip()
        # Split into parts and keep only first 3 words (title + firstname + lastname)
        parts = name.split()
        return ' '.join(parts[:2]) if parts else ''

    def extract_text_from_pdf(self, file_path):
        text = ""
        try:
            reader = PdfReader(file_path)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        except Exception as e:
            raise Exception(f"PDF reading error: {str(e)}")
        return text

    def extract_text_from_docx(self, file_path):
        text = ""
        try:
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
        except Exception as e:
            raise Exception(f"DOCX reading error: {str(e)}")
        return text

    def process_file(self, file_path):
        # Clear previous entries
        for item in self.tree.get_children():
            self.tree.delete(item)

        try:
            # Extract text based on file type
            if file_path.lower().endswith('.pdf'):
                text = self.extract_text_from_pdf(file_path)
            else:  # .docx
                text = self.extract_text_from_docx(file_path)

            # Thai name pattern with titles
            name_pattern = r'\d+\)\s*((?:นาย|นาง|นางสาว|ว่าที่[ก-์\s]+|ว่าที่พันตรี|ว่าที่ร้อยตรี)[ก-์\s]+[ก-์]+)'
            names = re.findall(name_pattern, text)
            
            # Clean names to keep only title + firstname + lastname
            names = [self.clean_thai_name(name) for name in names]
            
            # Count occurrences
            name_counts = Counter(names)
            
            # Store all names for filtering
            self.all_names = sorted(name_counts.items())
            
            # Display results
            for name, count in self.all_names:
                self.tree.insert('', tk.END, values=(name, count))

            if not self.all_names:
                messagebox.showwarning("คำเตือน", "ไม่พบรายชื่อในไฟล์")

        except Exception as e:
            messagebox.showerror("Error", f"เกิดข้อผิดพลาด: {str(e)}")

    def export_to_excel(self):
        if not self.all_names:
            messagebox.showwarning("คำเตือน", "ไม่มีข้อมูลที่จะส่งออก")
            return
            
        try:
            # Get save file location
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=f"name_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not file_path:
                return
                
            # Create workbook and select active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Name List"
            
            # Add headers
            ws['A1'] = 'ชื่อ'
            ws['B1'] = 'จำนวน'
            
            # Add data
            for row, (name, count) in enumerate(self.all_names, start=2):
                ws[f'A{row}'] = name
                ws[f'B{row}'] = count
            
            # Save workbook
            wb.save(file_path)
            messagebox.showinfo("สำเร็จ", "ส่งออกไฟล์ Excel เรียบร้อยแล้ว")
            
        except Exception as e:
            messagebox.showerror("Error", f"เกิดข้อผิดพลาดในการส่งออก: {str(e)}")

def main():
    root = tk.Tk()
    app = NameExtractorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
